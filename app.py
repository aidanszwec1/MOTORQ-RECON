#!/usr/bin/env python3
import streamlit as st
import pandas as pd
import tempfile
import os
import re
from typing import Dict
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

st.set_page_config(
    page_title="Motorq Reconciliation",
    page_icon="🚗",
    layout="wide"
)

# Motorq reconciliation functions (copied from motorq_recon.py)
VIN_COL_INTERNAL = "VIN"
DAYS_ENROLLED_COL_INTERNAL = "DAYS_ENROLLED"
VIN_COL_INVOICE = "VIN/fleetUnitId"
PRODUCT_COL_INVOICE = "Motorq product"
DAYS_ENROLLED_COL_INVOICE = "# Days enrolled"

def normalize_vin(value) -> str:
    if pd.isna(value):
        return ""
    s = str(value).strip().upper()
    # Remove spaces and non-alphanumeric just in case
    s = re.sub(r"[^A-Z0-9]", "", s)
    return s

def load_internal_days_enrolled_map(all_motorq_csv: str) -> Dict[str, str]:
    df = pd.read_csv(all_motorq_csv, dtype=str, low_memory=False)
    if VIN_COL_INTERNAL not in df.columns:
        raise ValueError(
            f"Expected column '{VIN_COL_INTERNAL}' in {all_motorq_csv} but it was not found. Found columns: {list(df.columns)}"
        )
    # Normalize VINs
    df[VIN_COL_INTERNAL] = df[VIN_COL_INTERNAL].map(normalize_vin)
    df = df[df[VIN_COL_INTERNAL].str.len() >= 11]

    # Ensure DAYS_ENROLLED column exists; if not, create empty
    if DAYS_ENROLLED_COL_INTERNAL not in df.columns:
        df[DAYS_ENROLLED_COL_INTERNAL] = ""

    # Aggregate days enrolled per VIN (max non-empty value, stored as string)
    def _max_days(series: pd.Series) -> str:
        values = []
        for x in series:
            if pd.isna(x):
                continue
            s = str(x).strip()
            if not s:
                continue
            try:
                values.append(float(s))
            except ValueError:
                continue
        if not values:
            return ""
        v = max(values)
        if float(v).is_integer():
            return str(int(v))
        return str(v)

    agg = (
        df.groupby(VIN_COL_INTERNAL, dropna=True)[DAYS_ENROLLED_COL_INTERNAL]
        .agg(_max_days)
        .reset_index()
    )
    return dict(zip(agg[VIN_COL_INTERNAL], agg[DAYS_ENROLLED_COL_INTERNAL]))

def _find_invoice_header_row(invoice_xlsx: str, sheet_name: str) -> int:
    preview = pd.read_excel(invoice_xlsx, sheet_name=sheet_name, header=None, engine="openpyxl", nrows=200)
    for i in range(len(preview)):
        row = preview.iloc[i].astype(str)
        if row.str.contains("VIN/fleetUnitId", case=False, na=False).any():
            return i
    raise ValueError(f"Could not locate invoice header row in sheet '{sheet_name}' for {invoice_xlsx}")

def load_invoice_detail(invoice_xlsx: str, sheet_name: str = "Invoice_Detail") -> pd.DataFrame:
    header_row = _find_invoice_header_row(invoice_xlsx, sheet_name)
    df = pd.read_excel(invoice_xlsx, sheet_name=sheet_name, header=header_row, dtype=str, engine="openpyxl")

    if VIN_COL_INVOICE not in df.columns:
        raise ValueError(
            f"Expected column '{VIN_COL_INVOICE}' in {invoice_xlsx} but it was not found. Found columns: {list(df.columns)}"
        )

    df[VIN_COL_INVOICE] = df[VIN_COL_INVOICE].map(normalize_vin)
    df = df[df[VIN_COL_INVOICE].str.len() >= 11]

    # Normalize optional columns
    if PRODUCT_COL_INVOICE not in df.columns:
        df[PRODUCT_COL_INVOICE] = ""
    if DAYS_ENROLLED_COL_INVOICE not in df.columns:
        df[DAYS_ENROLLED_COL_INVOICE] = ""

    # One row per VIN (aggregate unique products, max invoice days enrolled)
    def _unique_join(series: pd.Series) -> str:
        vals = []
        for x in series:
            if pd.isna(x):
                continue
            s = str(x).strip()
            if not s:
                continue
            vals.append(s)
        return "; ".join(sorted(set(vals)))

    def _max_days_str(series: pd.Series) -> str:
        values = []
        for x in series:
            if pd.isna(x):
                continue
            s = str(x).strip()
            if not s:
                continue
            try:
                values.append(float(s))
            except ValueError:
                continue
        if not values:
            return ""
        v = max(values)
        if float(v).is_integer():
            return str(int(v))
        return str(v)

    grouped = (
        df.groupby(VIN_COL_INVOICE, dropna=True)
        .agg(
            {
                PRODUCT_COL_INVOICE: _unique_join,
                DAYS_ENROLLED_COL_INVOICE: _max_days_str,
            }
        )
        .reset_index()
    )
    grouped = grouped.rename(
        columns={
            VIN_COL_INVOICE: "VIN",
            PRODUCT_COL_INVOICE: "MOTORQ_PRODUCT",
            DAYS_ENROLLED_COL_INVOICE: "INVOICE_DAYS_ENROLLED",
        }
    )
    grouped["MOTORQ_PRODUCT"] = grouped["MOTORQ_PRODUCT"].fillna("").astype(str)
    grouped["INVOICE_DAYS_ENROLLED"] = grouped["INVOICE_DAYS_ENROLLED"].fillna("").astype(str)
    return grouped

def _style_headers(ws):
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border

def _autofit_columns(ws):
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, val in enumerate(row, start=1):
            val_str = "" if val is None else str(val)
            widths[i] = max(widths.get(i, 0), len(val_str))
    for i, w in widths.items():
        ws.column_dimensions[chr(64 + i)].width = min(max(w + 2, 12), 60)

def write_recon(
    output_xlsx: str,
    internal_days_map: Dict[str, str],
    invoice_detail: pd.DataFrame,
    internal_label: str,
    invoice_label: str,
):
    internal_vins = set(internal_days_map.keys())
    invoice_vins = set(invoice_detail["VIN"].dropna().tolist())
    on_invoice_not_internal = invoice_vins - internal_vins
    in_internal_not_invoice = internal_vins - invoice_vins
    matches = invoice_vins & internal_vins

    invoice_by_vin = invoice_detail.set_index("VIN")

    df_on_inv = pd.DataFrame(sorted(on_invoice_not_internal), columns=["VIN"])
    df_on_inv["MOTORQ_PRODUCT"] = df_on_inv["VIN"].map(lambda v: invoice_by_vin.loc[v, "MOTORQ_PRODUCT"] if v in invoice_by_vin.index else "")
    df_on_inv["INVOICE_DAYS_ENROLLED"] = df_on_inv["VIN"].map(
        lambda v: invoice_by_vin.loc[v, "INVOICE_DAYS_ENROLLED"] if v in invoice_by_vin.index else ""
    )
    df_on_inv["INTERNAL_DAYS_ENROLLED"] = ""

    df_in_internal = pd.DataFrame(sorted(in_internal_not_invoice), columns=["VIN"])
    df_in_internal["MOTORQ_PRODUCT"] = ""
    df_in_internal["INVOICE_DAYS_ENROLLED"] = ""
    df_in_internal["INTERNAL_DAYS_ENROLLED"] = df_in_internal["VIN"].map(internal_days_map)

    df_matches = pd.DataFrame(sorted(matches), columns=["VIN"])
    df_matches["MOTORQ_PRODUCT"] = df_matches["VIN"].map(
        lambda v: invoice_by_vin.loc[v, "MOTORQ_PRODUCT"] if v in invoice_by_vin.index else ""
    )
    df_matches["INVOICE_DAYS_ENROLLED"] = df_matches["VIN"].map(
        lambda v: invoice_by_vin.loc[v, "INVOICE_DAYS_ENROLLED"] if v in invoice_by_vin.index else ""
    )
    df_matches["INTERNAL_DAYS_ENROLLED"] = df_matches["VIN"].map(internal_days_map)

    df_on_inv = df_on_inv.fillna("")
    df_in_internal = df_in_internal.fillna("")
    df_matches = df_matches.fillna("")

    # Summary
    total_internal = len(internal_vins)
    total_invoice = len(invoice_vins)
    matched = len(matches)
    missing_from_internal = len(on_invoice_not_internal)

    df_summary = pd.DataFrame(
        [
            [f"Total VINs in {internal_label}", total_internal],
            [f"Total VINs on {invoice_label}", total_invoice],
            [f"Invoice VINs found in {internal_label}", matched],
            [f"Invoice VINs missing in {internal_label}", missing_from_internal],
        ],
        columns=["Metric", "Value"],
    )

    df_by_product = (
        invoice_detail[["VIN", "MOTORQ_PRODUCT"]]
        .dropna(subset=["VIN"])
        .loc[lambda d: d["MOTORQ_PRODUCT"].fillna("").astype(str).str.strip() != ""]
        .groupby("MOTORQ_PRODUCT", dropna=True)
        .agg(VIN_COUNT=("VIN", "nunique"))
        .reset_index()
        .sort_values(["VIN_COUNT", "MOTORQ_PRODUCT"], ascending=[False, True])
    )

    # Write to Excel with styles
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        df_summary.to_excel(writer, index=False, sheet_name="Summary")
        df_by_product.to_excel(writer, index=False, sheet_name="ByProduct")
        df_on_inv.to_excel(writer, index=False, sheet_name="OnInvoice_NotInInternal")
        df_in_internal.to_excel(writer, index=False, sheet_name="InInternal_NotOnInvoice")
        df_matches.to_excel(writer, index=False, sheet_name="Matches")

        wb = writer.book
        for name in ["Summary", "ByProduct", "OnInvoice_NotInInternal", "InInternal_NotOnInvoice", "Matches"]:
            ws = wb[name]
            _style_headers(ws)
            _autofit_columns(ws)
            ws.freeze_panes = "A2"

    return df_summary

def run_motorq_reconciliation(all_motorq_csv: str, invoice_xlsx: str, output_xlsx: str):
    internal_days_map = load_internal_days_enrolled_map(all_motorq_csv)
    invoice_detail = load_invoice_detail(invoice_xlsx)
    os.makedirs(os.path.dirname(output_xlsx) or ".", exist_ok=True)
    return write_recon(
        output_xlsx,
        internal_days_map,
        invoice_detail,
        internal_label="ALL_MOTORQ.csv",
        invoice_label="Motorq Invoice",
    )

# Streamlit UI
st.title("🚗 Motorq Invoice Reconciliation")
st.markdown("Upload the internal Motorq CSV and the invoice XLSX, then run the reconciliation.")

# Create two columns for file uploads
col1, col2 = st.columns(2)

with col1:
    st.subheader("Internal Motorq Data (CSV)")
    internal_file = st.file_uploader(
        "Upload ALL_MOTORQ CSV file",
        type=['csv'],
        help="Expected format: CSV with VIN and DAYS_ENROLLED columns",
        key="internal"
    )
    if internal_file:
        st.success(f"✅ {internal_file.name} uploaded ({internal_file.size} bytes)")

with col2:
    st.subheader("Motorq Invoice (XLSX)")
    invoice_file = st.file_uploader(
        "Upload Motorq invoice XLSX file", 
        type=['xlsx'],
        help="Expected format: XLSX with 'Invoice_Detail' sheet containing VIN/fleetUnitId column",
        key="invoice"
    )
    if invoice_file:
        st.success(f"✅ {invoice_file.name} uploaded ({invoice_file.size} bytes)")

# Run reconciliation button
if st.button("🔄 Run Reconciliation", type="primary", use_container_width=True):
    if not internal_file or not invoice_file:
        st.error("❌ Please upload both files before running the reconciliation.")
    else:
        with st.spinner("Running reconciliation..."):
            try:
                # Create temporary files
                with tempfile.TemporaryDirectory() as temp_dir:
                    # Save uploaded files to temporary directory
                    internal_path = os.path.join(temp_dir, "internal_data.csv")
                    invoice_path = os.path.join(temp_dir, "invoice_data.xlsx")
                    output_path = os.path.join(temp_dir, "MOTORQ_RECON_OUTPUT.xlsx")
                    
                    # Write uploaded files
                    with open(internal_path, "wb") as f:
                        f.write(internal_file.getbuffer())
                    
                    with open(invoice_path, "wb") as f:
                        f.write(invoice_file.getbuffer())
                    
                    # Run the reconciliation
                    summary_df = run_motorq_reconciliation(internal_path, invoice_path, output_path)
                    
                    # Read the output file for download
                    with open(output_path, "rb") as f:
                        output_data = f.read()
                    
                    # Display success message and download button
                    st.success("✅ Reconciliation completed successfully!")
                    
                    # Create download button
                    st.download_button(
                        label="📥 Download Reconciliation Results",
                        data=output_data,
                        file_name=f"MOTORQ_RECON_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    
                    # Show preview of results
                    st.subheader("📊 Reconciliation Summary")
                    
                    # Display metrics in columns
                    metrics_col1, metrics_col2 = st.columns(2)
                    
                    for idx, row in summary_df.iterrows():
                        metric_name = row['Metric']
                        metric_value = row['Value']
                        
                        if idx % 2 == 0:
                            with metrics_col1:
                                st.metric(metric_name, metric_value)
                        else:
                            with metrics_col2:
                                st.metric(metric_name, metric_value)
                    
                    # Show the summary table
                    st.dataframe(summary_df, use_container_width=True)
                    
            except Exception as e:
                st.error(f"❌ Error during reconciliation: {str(e)}")
                st.error("Please check that your files are in the correct format and try again.")

# Instructions section
with st.expander("📋 Instructions", expanded=False):
    st.markdown("""
    ### How to use this tool:
    
    1. **Internal Motorq Data (CSV)**: Upload your ALL_MOTORQ CSV file containing VIN and DAYS_ENROLLED columns
    2. **Motorq Invoice (XLSX)**: Upload the Motorq invoice Excel file with an 'Invoice_Detail' sheet
    3. Click **Run Reconciliation** to process the files
    4. Download the results Excel file when processing is complete
    
    ### Output file contains:
    - **Summary**: Overall reconciliation metrics
    - **ByProduct**: VIN counts by Motorq product
    - **OnInvoice_NotInInternal**: VINs on invoice but missing in internal data
    - **InInternal_NotOnInvoice**: VINs in internal data but missing on invoice
    - **Matches**: VINs that match between both datasets
    
    ### File format requirements:
    - **CSV file**: Must contain 'VIN' and 'DAYS_ENROLLED' columns
    - **XLSX file**: Must contain 'Invoice_Detail' sheet with 'VIN/fleetUnitId' column
    """)

# Footer
st.markdown("---")
st.markdown("*Motorq Invoice Reconciliation Tool - Automated reconciliation between internal data and invoices*")
